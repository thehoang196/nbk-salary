"""
NBK Salary - Export Service

Generates Excel exports in Misa-compatible format and salary summary reports.
"""
from io import BytesIO
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from sqlalchemy.orm import Session

from app.models.bang_luong import DlBangLuong
from app.models.nhan_vien import NhanVien


class MisaExporter:
    """Generates Misa-compatible Excel export for salary data."""
    
    def __init__(self, db: Session):
        self.db = db

    def export_salary_to_misa(self, thang: int, nam: int) -> BytesIO:
        """Export salary data in Misa bulk import format.
        
        Returns BytesIO containing the .xlsx file.
        """
        records = self.db.query(DlBangLuong).filter(
            DlBangLuong.thang == thang,
            DlBangLuong.nam == nam,
        ).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Lương T{thang}/{nam}"
        
        # Header row
        headers = [
            "STT", "Mã NV", "Họ tên", "CCCD", "Nhóm NV",
            "Lương khoán", "Tiền giảng dạy", "Hỗ trợ",
            "Tổng thu nhập", "BH + CĐ", "Thuế TNCN",
            "Tổng giảm trừ", "Thực lĩnh",
        ]
        ws.append(headers)
        
        # Bold header
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Data rows
        for idx, r in enumerate(records, 1):
            nv = self.db.query(NhanVien).filter(NhanVien.id == r.nhan_vien_id).first()
            muc_ii = r.muc_ii_json or {}
            muc_iv = r.muc_iv_json or {}
            
            ws.append([
                idx,
                nv.ma_nv if nv else "",
                nv.ho_ten if nv else "",
                nv.cccd if nv else "",
                nv.nhom_nv.value if nv else "",
                muc_ii.get("luong_khoan_prorated", 0),
                muc_ii.get("tien_giang_day", 0),
                muc_ii.get("ho_tro", 0),
                muc_ii.get("tong", 0),
                muc_iv.get("bh_cd", 0),
                muc_iv.get("thue_tncn", 0),
                muc_iv.get("tong", 0),
                int(r.muc_vi_thuc_linh or 0),
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_salary_summary(self, thang: int, nam: int) -> BytesIO:
        """Export monthly salary summary report as Excel."""
        records = self.db.query(DlBangLuong).filter(
            DlBangLuong.thang == thang,
            DlBangLuong.nam == nam,
        ).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Tổng hợp lương T{thang}/{nam}"
        
        ws.append(["BẢNG TỔNG HỢP LƯƠNG", "", f"Tháng {thang}/{nam}"])
        ws.append([])
        ws.append(["STT", "Mã NV", "Họ tên", "Nhóm", "Trạng thái", "Thực lĩnh"])
        
        total = 0
        for idx, r in enumerate(records, 1):
            nv = self.db.query(NhanVien).filter(NhanVien.id == r.nhan_vien_id).first()
            thuc_linh = int(r.muc_vi_thuc_linh or 0)
            ws.append([
                idx, nv.ma_nv if nv else "", nv.ho_ten if nv else "",
                nv.nhom_nv.value if nv else "", r.trang_thai.value, thuc_linh
            ])
            total += thuc_linh
        
        ws.append([])
        ws.append(["", "", "", "", "TỔNG:", total])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
