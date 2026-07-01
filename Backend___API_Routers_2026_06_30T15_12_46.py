
# ============================================================
# FILE: backend/app/services/export_service.py
# Xuất báo cáo: Phiếu lương PDF, Form Misa Excel
# ============================================================

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from datetime import date
from typing import List
from sqlalchemy.orm import Session
from app.models.luong import BangLuong
from app.models.nhan_vien import NhanVien


class ExportService:
    """Service xuất báo cáo lương"""
    
    def __init__(self, db: Session):
        self.db = db
    
    # ========================================================
    # Xuất Form Misa (Excel) - để import vào phần mềm Misa
    # ========================================================
    def export_misa_format(self, thang: int, nam: int) -> BytesIO:
        """
        Xuất bảng lương theo format Misa
        Columns: Mã NV, Tên, Phòng ban, Lương CB, Phụ cấp, 
                 Tổng TN, BHXH, BHYT, BHTN, Thuế, Thực lĩnh
        """
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Luong_T{thang}_{nam}"
        
        # Header theo format Misa
        headers = [
            "Mã nhân viên", "Họ và tên", "Phòng ban", "Chức danh",
            "Số ngày công", "Lương cơ bản", "Lương chính",
            "Lương dạy", "Lương hiệu quả",
            "Phụ cấp CV", "Phụ cấp TN", "Phụ cấp thâm niên",
            "Phụ cấp ăn trưa", "Phụ cấp khác",
            "Tăng ca", "Thưởng", "Bổ sung", "Giảm lương",
            "Tổng thu nhập",
            "BHXH (8%)", "BHYT (1.5%)", "BHTN (1%)",
            "Thuế TNCN",
            "Tổng khấu trừ",
            "Thực lĩnh",
            "Số tài khoản", "Ngân hàng"
        ]
        
        # Style header
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        
        # Data
        bang_luongs = self.db.query(BangLuong).filter(
            BangLuong.thang == thang,
            BangLuong.nam == nam,
            BangLuong.trang_thai == "approved"
        ).all()
        
        for row_idx, bl in enumerate(bang_luongs, 2):
            nv = self.db.query(NhanVien).filter(NhanVien.id == bl.nhan_vien_id).first()
            
            ws.cell(row=row_idx, column=1, value=nv.ma_nv)
            ws.cell(row=row_idx, column=2, value=nv.ho_ten)
            ws.cell(row=row_idx, column=3, value=nv.don_vi.ten_don_vi if nv.don_vi else "")
            ws.cell(row=row_idx, column=4, value=nv.chuc_danh.ten_chuc_danh if nv.chuc_danh else "")
            ws.cell(row=row_idx, column=5, value="")  # Số ngày công - from tong_hop_cong
            ws.cell(row=row_idx, column=6, value=bl.luong_chuc_danh)
            ws.cell(row=row_idx, column=7, value=bl.luong_chinh)
            ws.cell(row=row_idx, column=8, value=bl.luong_day)
            ws.cell(row=row_idx, column=9, value=bl.luong_hieu_qua)
            ws.cell(row=row_idx, column=10, value=bl.phu_cap_1)
            ws.cell(row=row_idx, column=11, value=bl.phu_cap_2)
            ws.cell(row=row_idx, column=12, value=bl.phu_cap_3)
            ws.cell(row=row_idx, column=13, value=bl.phu_cap_4)
            ws.cell(row=row_idx, column=14, value=bl.phu_cap_5)
            ws.cell(row=row_idx, column=15, value=bl.tang_ca)
            ws.cell(row=row_idx, column=16, value=bl.thuong)
            ws.cell(row=row_idx, column=17, value=bl.bo_sung)
            ws.cell(row=row_idx, column=18, value=bl.giam_luong)
            ws.cell(row=row_idx, column=19, value=bl.tong_thu_nhap)
            ws.cell(row=row_idx, column=20, value=bl.bhxh)
            ws.cell(row=row_idx, column=21, value=bl.bhyt)
            ws.cell(row=row_idx, column=22, value=bl.bhtn)
            ws.cell(row=row_idx, column=23, value=bl.thue_tncn)
            ws.cell(row=row_idx, column=24, value=bl.tong_khau_tru)
            ws.cell(row=row_idx, column=25, value=bl.thuc_linh)
            ws.cell(row=row_idx, column=26, value=nv.so_tai_khoan or "")
            ws.cell(row=row_idx, column=27, value=nv.ngan_hang or "")
            
            # Format số tiền
            for col in range(6, 26):
                ws.cell(row=row_idx, column=col).number_format = '#,##0'
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    # ========================================================
    # Xuất Báo cáo tổng hợp lương
    # ========================================================
    def export_tong_hop_luong(self, thang: int, nam: int) -> BytesIO:
        """Xuất báo cáo tổng hợp lương theo phòng ban"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "TongHopLuong"
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"BÁO CÁO TỔNG HỢP LƯƠNG THÁNG {thang}/{nam}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Headers
        headers = ["STT", "Phòng ban", "Số NV", "Tổng lương chính", 
                   "Tổng lương dạy", "Tổng phụ cấp", "Tổng khấu trừ", "Tổng thực lĩnh"]
        
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
            ws.cell(row=3, column=col).font = Font(bold=True)
        
        # Aggregate by department
        from sqlalchemy import func
        from app.models.danh_muc import DonVi
        
        results = self.db.query(
            DonVi.ten_don_vi,
            func.count(BangLuong.id),
            func.sum(BangLuong.luong_chinh),
            func.sum(BangLuong.luong_day),
            func.sum(BangLuong.phu_cap_1 + BangLuong.phu_cap_2 + BangLuong.phu_cap_3 + 
                    BangLuong.phu_cap_4 + BangLuong.phu_cap_5),
            func.sum(BangLuong.tong_khau_tru),
            func.sum(BangLuong.thuc_linh)
        ).join(NhanVien, BangLuong.nhan_vien_id == NhanVien.id
        ).join(DonVi, NhanVien.don_vi_id == DonVi.id
        ).filter(
            BangLuong.thang == thang,
            BangLuong.nam == nam
        ).group_by(DonVi.ten_don_vi).all()
        
        for idx, row in enumerate(results, 1):
            ws.cell(row=idx+3, column=1, value=idx)
            ws.cell(row=idx+3, column=2, value=row[0])
            ws.cell(row=idx+3, column=3, value=row[1])
            for col in range(4, 9):
                ws.cell(row=idx+3, column=col, value=row[col-2] or 0)
                ws.cell(row=idx+3, column=col).number_format = '#,##0'
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

